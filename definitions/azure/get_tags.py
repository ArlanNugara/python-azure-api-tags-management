import requests
import json
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import sys

# Declare the CSV header here
subscription_header = ['Subscription Name', 'Subscription ID','Scope', 'Lock Details', 'Original Tags', 'Replace Tags', 'Merge Tags', 'Delete Tags']
rg_header = ['Subscription Name', 'Subscription ID', 'Resource Group Name', 'Resource Type', 'Scope', 'Location', 'Lock Details', 'Original Tags', 'Replace Tags', 'Merge Tags', 'Delete Tags']
resource_header = ['Subscription Name', 'Subscription ID', 'Resource Group Name', 'Resource Name', 'Resource Type', 'Scope', 'Location', 'Lock Details', 'Original Tags', 'Replace Tags', 'Merge Tags', 'Delete Tags']

# Write Subscription Header
subscription_excel_header = pd.DataFrame(columns = subscription_header)
subscription_excel_writer = pd.ExcelWriter('Reports.xlsx', engine='openpyxl')
subscription_excel_header.to_excel(subscription_excel_writer, sheet_name='Subscription', index=False)
subscription_excel_writer.save()
print("Successfuly created Subscription Worksheet")

# Write Resource Group Header
resource_group_excel_header = pd.DataFrame(columns = rg_header)
resource_group_excel_writer = pd.ExcelWriter('Reports.xlsx', engine='openpyxl', mode='a')
resource_group_excel_header.to_excel(resource_group_excel_writer, sheet_name='ResourceGroup', index=False)
resource_group_excel_writer.save()
print("Successfuly created Resource Group Worksheet")

# Write Resource Header
resource_excel_header = pd.DataFrame(columns = resource_header)
resource_excel_writer = pd.ExcelWriter('Reports.xlsx', engine='openpyxl', mode='a')
resource_excel_header.to_excel(resource_excel_writer, sheet_name='Resource', index=False)
resource_excel_writer.save()
print("Successfully created the Resource Worksheet")

# Start API calls
def get_all_tags(main_url, subscription_id, header):
    # Get Subscription Details
    print("Getting Details for Subscription ID - "+subscription_id+"")
    get_subscription_details = requests.get(url = ""+main_url+"/subscriptions/"+subscription_id+"?api-version=2020-01-01", headers = header)
    subscription_response_to_json = get_subscription_details.json()
    print(subscription_response_to_json)
    if get_subscription_details.status_code == 200 or get_subscription_details.status_code == 204:
        print("API Response Received for Subscription ID - "+subscription_id+"")
        subscription_lock_details = requests.get(url = ""+main_url+""+subscription_response_to_json["id"]+"/providers/Microsoft.Authorization/locks?api-version=2016-09-01", headers = header)
        subscription_lock_resonse_to_json = subscription_lock_details.json()
        if subscription_lock_details.status_code == 200 or subscription_lock_details.status_code == 204:
            if "value" in subscription_lock_resonse_to_json:
                subscription_lock_data = []
                for locks in subscription_lock_resonse_to_json['value']:
                    lock_json = {"type": locks['properties']['level'], "name": locks['name'], "id": locks['id']}
                    subscription_lock_data.append(lock_json)
                subscription_lock_excel_value = {"locks": subscription_lock_data}
            else:
                subscription_lock_excel_value = "No Locks"
        else:
            subscription_lock_excel_value = "Error in API response"
        
        if "tags" in subscription_response_to_json:
            subscription_excel_data = [[subscription_response_to_json["displayName"], subscription_response_to_json["subscriptionId"], subscription_response_to_json["id"], json.dumps(subscription_lock_excel_value, indent = 2), json.dumps(subscription_response_to_json["tags"], indent=2), None, None, None]]
        else:
            subscription_excel_data = [[subscription_response_to_json["displayName"], subscription_response_to_json["subscriptionId"], subscription_response_to_json["id"], json.dumps(subscription_lock_excel_value, indent = 2), "No Tags", None, None, None]]
        subscription_workbook = load_workbook(filename = "Reports.xlsx")
        subscription_worksheet = subscription_workbook["Subscription"]
        for subscription_rows in dataframe_to_rows(pd.DataFrame(subscription_excel_data), index=False, header=False):
            subscription_worksheet.append(subscription_rows)
        subscription_worksheet.auto_filter.ref = subscription_worksheet.dimensions
        subscription_workbook.save("Reports.xlsx")
        subscription_workbook.close()
        print("Finished writing Subscription Details to Excel File")
    else:
        print("Error in API response for ", subscription_id)
        sys.exit(1)

    # Get list of resource groups
    print("Getting Resource Groups Details for Subscription ID - "+subscription_id+"")
    get_rg_details = requests.get(url = ""+main_url+"/subscriptions/"+subscription_id+"/resourcegroups?api-version=2021-04-01", headers = header)
    rg_response_to_json = get_rg_details.json()
    if get_rg_details.status_code == 200 or get_rg_details.status_code == 204:
        if "value" in rg_response_to_json:
            rg_lock_data = []
            rg_excel_entry_list = []
            for rg in rg_response_to_json["value"]:
                rg_lock_details = requests.get(url = ""+main_url+""+rg["id"]+"/providers/Microsoft.Authorization/locks?api-version=2016-09-01", headers = header)
                rg_lock_resonse_to_json = rg_lock_details.json()
                if rg_lock_details.status_code == 200 or rg_lock_details.status_code == 204:
                    if "value" in rg_lock_resonse_to_json:
                        for locks in rg_lock_resonse_to_json["value"]:
                            lock_json = {"type": locks['properties']['level'], "name": locks['name'], "id": locks['id']}
                            rg_lock_data.append(lock_json)
                        rg_lock_excel_value = json.dumps({"locks": rg_lock_data}, indent=2)
                    else:
                        rg_lock_excel_value = "No Locks"
                else:
                    rg_lock_excel_value = "Error in API response"
                if "tags" in rg:
                    rg_excel_data = [subscription_response_to_json["displayName"], subscription_response_to_json["subscriptionId"], rg["name"], rg["type"], rg["id"], rg["location"], rg_lock_excel_value, json.dumps(rg["tags"], indent=2), None, None, None]
                else:
                    rg_excel_data = [subscription_response_to_json["displayName"], subscription_response_to_json["subscriptionId"], rg["name"], rg["type"], rg["id"], rg["location"], rg_lock_excel_value, "No Tags", None, None, None]
                rg_excel_entry_list.append(rg_excel_data)
            print("Resource Group API Response Received for Subscription ID - "+subscription_id+"")
            resource_group_workbook = load_workbook(filename = "Reports.xlsx")
            resource_group_worksheet = resource_group_workbook["ResourceGroup"]
            for resource_group_rows in dataframe_to_rows(pd.DataFrame(rg_excel_entry_list), index=False, header=False):
                resource_group_worksheet.append(resource_group_rows)
            resource_group_worksheet.auto_filter.ref = resource_group_worksheet.dimensions
            resource_group_workbook.save("Reports.xlsx")
            resource_group_workbook.close()
            print("Finished writing Resource Group Details to Excel File")
        else:
            print("No Resource Group present in ", subscription_id)
            pass
    else:
        print("Error in Resource group API response for ", subscription_id)
        sys.exit(1)

    # Get list of resources
    for rg in rg_response_to_json["value"]:
        print("Getting Resources Details for Resource Group - "+rg["id"].split('/')[4]+"")
        get_resource_details = requests.get(url = ""+main_url+""+rg["id"]+"/resources?api-version=2021-04-01", headers = header)
        resource_response_to_json = get_resource_details.json()
        if get_resource_details.status_code == 200 or get_resource_details.status_code == 204:
            if "value" in resource_response_to_json:
                resource_lock_data = []
                resource_excel_entry_list = []
                for resource in resource_response_to_json["value"]:
                    resource_lock_details = requests.get(url = ""+main_url+""+resource["id"]+"/providers/Microsoft.Authorization/locks?api-version=2016-09-01", headers = header)
                    resource_lock_resonse_to_json = resource_lock_details.json()
                    if resource_lock_details.status_code == 200 or resource_lock_details.status_code == 204:
                        if "value" in resource_lock_resonse_to_json:
                            for locks in resource_lock_resonse_to_json["value"]:
                                lock_json = {"type": locks['properties']['level'], "name": locks['name'], "id": locks['id']}
                                resource_lock_data.append(lock_json)
                            resource_lock_excel_value = json.dumps({"locks": resource_lock_data}, indent=2)
                        else:
                            resource_lock_excel_value = "No Locks"
                    else:
                        resource_lock_excel_value = "Error in API response"
                    if "tags" in resource:
                        rg_name_split = resource["id"].split('/')
                        rg_name = rg_name_split[4]
                        resource_excel_data = [subscription_response_to_json["displayName"], subscription_response_to_json["subscriptionId"], rg_name, resource["name"], resource["type"], resource["id"], resource["location"], resource_lock_excel_value, json.dumps(resource["tags"], indent=2), None, None, None]
                    else:
                        rg_name_split = resource["id"].split('/')
                        rg_name = rg_name_split[4]
                        resource_excel_data = [subscription_response_to_json["displayName"], subscription_response_to_json["subscriptionId"], rg_name, resource["name"], resource["type"], resource["id"], resource["location"], resource_lock_excel_value, "No Tags", None, None, None]
                    resource_excel_entry_list.append(resource_excel_data)
                resource_workbook = load_workbook(filename = "Reports.xlsx")
                resource_worksheet = resource_workbook["Resource"]
                for resource_rows in dataframe_to_rows(pd.DataFrame(resource_excel_entry_list), index=False, header=False):
                    resource_worksheet.append(resource_rows)
                resource_worksheet.auto_filter.ref = resource_worksheet.dimensions
                resource_workbook.save("Reports.xlsx")
                resource_workbook.close()
            else:
                print("No Resource Present in ", rg_name)
                pass
        else:
            print("Error in Resource API response for ", rg_name)
            sys.exit(1)